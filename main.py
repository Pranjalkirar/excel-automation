import os
from openpyxl import load_workbook

def column_letter_to_index(letter):
    return ord(letter.upper()) - ord('A') + 1

def process_workbook(file_in, file_out, source_col, target_col, expression):
    wb = load_workbook(file_in)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, source_col)
        if cell.value is not None:
            x = cell.value  # ✅ x is now used inside eval()
            try:
                result = eval(expression)
                sheet.cell(row, target_col).value = result
            except Exception as e:
                print(f"⚠️ Error at row {row} in file {file_in}: {e}")

    wb.save(file_out)

# User Inputs
print("🔧 Simple Excel Modifier")

source_col_letter = input("Enter column letter to modify (e.g., C): ").strip().upper()
target_col_letter = input("Enter column letter to write results (e.g., D): ").strip().upper()
user_expression = input("Enter math expression using 'x' (e.g., x * 0.9, x + 100): ").strip()

source_col_index = column_letter_to_index(source_col_letter)
target_col_index = column_letter_to_index(target_col_letter)

# Folder Setup
input_folder = 'input'
output_folder = 'output'
os.makedirs(output_folder, exist_ok=True)

# Loop through files
for file_name in os.listdir(input_folder):
    if file_name.endswith('.xlsx') and not file_name.startswith('~$'):
        full_in_path = os.path.join(input_folder, file_name)
        full_out_path = os.path.join(output_folder, file_name.replace('.xlsx', '_corrected.xlsx'))

        print(f"Processing {file_name}...")
        process_workbook(full_in_path, full_out_path, source_col_index, target_col_index, user_expression)

print(" Done! Modified files are saved in the 'output/' folder.")
